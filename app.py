from fastapi import FastAPI, Request,Form,Response,File,UploadFile,HTTPException
from fastapi.staticfiles import StaticFiles
from starlette.types import Scope
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse,RedirectResponse, HTMLResponse
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from PIL import Image
# from openpyxl.drawing.image import Image
import string
import random
import hashlib
import os
import zipfile
import brotli
from brotli_asgi import BrotliMiddleware

# Алфавит для создания случайных названий картинок
alphabet = string.digits + string.ascii_lowercase


templates=Jinja2Templates(directory='templates')
app = FastAPI()

client = MongoClient("db", 27017)
# client.drop_database('test_database')
# client.drop_database('password_database')
# client.drop_database('session_database')
db = client.test_database
dbPassword = client.password_database
dbSession = client.session_database
collection = client.test_collection
collectionPassword = client.password_collection
posts = db.posts
postsPassword = dbPassword.posts
postsSession = dbSession.posts


app.add_middleware(BrotliMiddleware)

class CacheControlledStaticFiles(StaticFiles):
    async def get_response(self, path: str, scope: Scope) -> Response:
        response = await super().get_response(path, scope)
        response.headers["Cache-Control"] = "public, max-age=31536000"
        return response

app.mount("/static", CacheControlledStaticFiles(directory="static"), name="static")

@app.get('/')
async def welcome(request:Request):
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postsSession.find_one({"Session": session_value})
    if(res==None):
        return RedirectResponse(url="/auth")
    return templates.TemplateResponse(name='index.html',context={'request':request})

@app.get('/auth')
async def welcome(request:Request) :
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postsSession.find_one({"Session": session_value})
    if(res!=None):
        return RedirectResponse(url="/")
    return templates.TemplateResponse(name='auth.html',context={'request':request})


@app.get('/register')
async def welcome(request:Request) :
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postsSession.find_one({"Session": session_value})
    if(res!=None):
        return RedirectResponse(url="/")
    return templates.TemplateResponse(name='register.html',context={'request':request})

async def new_delete_file(name: str):
    path = "static/images/"
    path+=name
    try:
        if(path!="static/images/"):
            os.remove(path)
    except FileNotFoundError:
        pass
    return {"message": "Файл успешно удалён"}


@app.put('/data')
async def upload(request: Request):
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postsSession.find_one({"Session": session_value})
    if(res==None):
        raise HTTPException(status_code=401, detail="Аккаунт не найден")

    form_data = await request.form()
    print(f"ПОЛУЧЕННЫЕ ДАННЫЕ{form_data}")
    id=form_data.get("id")
    photo_device = form_data.get("photo_device")
    photo_serial_number_device = form_data.get("photo_serial_number_device")
    photo_ITAM_device = form_data.get("photo_ITAM_device")
    print(f"НАЗВАНИЕ ФАЙЛА: {photo_device}")

    check = posts.find_one({"user_id":res["id"],"id": int(id)})
    if(check==None):
        raise HTTPException(status_code=404, detail="Запись не найдена")
    filter = {'user_id':res["id"],'id': int(id)}

    posts.update_many(filter, {'$set': {'type_device': form_data.get("type_device")}})
    posts.update_many(filter, {'$set': {'model_device': form_data.get("model_device")}})
    posts.update_many(filter, {'$set': {'serial_number': form_data.get("serial_number")}})
    posts.update_many(filter, {'$set': {'ITAM_device': form_data.get("ITAM_device")}})
    delete = posts.find_one({'user_id':res["id"],'id': int(id)})
    if(photo_device.filename!=""):
        await new_delete_file(delete['photo_device'])
        name_photo_device = await newUpload(photo_device)
        posts.update_many(filter, {'$set': {'photo_device': name_photo_device}})
    if(photo_serial_number_device.filename!=""):
        await new_delete_file(delete['photo_serial_number_device'])
        name_photo_serial_number_device = await newUpload(photo_serial_number_device)
        posts.update_many(filter, {'$set': {'photo_serial_number_device': name_photo_serial_number_device}})
    if(photo_ITAM_device.filename!=""):
        await new_delete_file(delete['photo_ITAM_device'])
        name_photo_ITAM_device = await newUpload(photo_ITAM_device)
        posts.update_many(filter, {'$set': {'photo_ITAM_device': name_photo_ITAM_device}})

    return {"message": "Запись успешно изменена"}

@app.get("/data")
async def read_data(request: Request):
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postsSession.find_one({"Session": session_value})
    if(res==None):
        raise HTTPException(status_code=401, detail="Аккаунт не найден")
    dataSend = posts.find({'user_id': res["id"]})
    posts_list = []
    for post in dataSend:
        post_dict = dict(post)
        del post_dict['_id']
        posts_list.append(post_dict)
    return posts_list

@app.delete('/data/{numDelete}')
async def delete(request: Request,numDelete: str):
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postsSession.find_one({"Session": session_value})
    if(res==None):
        raise HTTPException(status_code=401, detail="Аккаунт не найден")
    count = posts.count_documents({})
    file = posts.find_one({"user_id":res["id"],"id": int(numDelete)})
    if(file==None):
        raise HTTPException(status_code=404, detail="Запись не найдена")
    path = "static/images/"
    fil1 = path+file["photo_device"]
    fil2 = path+file["photo_serial_number_device"]
    fil3 = path+file["photo_ITAM_device"]
    try:
        if(fil1!="static/images/"):
            os.remove(fil1)
        if(fil2!="static/images/"):
            os.remove(fil2)
        if(fil3!="static/images/"):
            os.remove(fil3)
    except FileNotFoundError:
        pass
    filterDelete = {'user_id':res["id"],'id': int(numDelete)}
    posts.delete_one(filterDelete)
    for i in range(int(numDelete)+1,count+1):
        filter = {'user_id':res["id"],'id': i}
        result = posts.update_one(filter, {'$set': {'id': i-1}})
    return {"message": "Запись успешно удалена"}

async def newUpload(file: UploadFile):
    newName = "".join([alphabet[random.randint(0, len(alphabet) -1)] for _ in range(0, 60)])
    exp=f".{file.filename.rsplit('.', 1)[1]}"
    tempName = newName + exp
    file.filename = tempName
    newName += ".webp"

    with open(f"static/images/{file.filename}", "wb") as f:
            f.write(await file.read())

    img = Image.open(f"static/images/{file.filename}")
    img.save(os.path.join("static/images", newName), format="webp")

    return newName

@app.post('/form')
async def sendForm(request: Request):
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postsSession.find_one({"Session": session_value})
    if(res==None):
        raise HTTPException(status_code=401, detail="Аккаунт не найден")
    form_data = await request.form()
    n=posts.count_documents({})
    newN = posts.find({'user_id': res["id"] })
    count = int(len(list(newN)))+1
    photo_device = form_data.get("photo_device")
    photo_serial_number_device = form_data.get("photo_serial_number_device")
    photo_ITAM_device = form_data.get("photo_ITAM_device")
    name_photo_device = await newUpload(photo_device)
    name_photo_serial_number_device = await newUpload(photo_serial_number_device)
    name_photo_ITAM_device = await newUpload(photo_ITAM_device)
    if( posts.count_documents({}) == n):
        post = {
            "user_id": res["id"],
            "id": count,
            "type_device": form_data.get("type_device"),
            "model_device": form_data.get("model_device"),
            "serial_number": form_data.get("serial_number"),
            "ITAM_device": form_data.get("ITAM_device"),
            "photo_device": name_photo_device,
            "photo_serial_number_device": name_photo_serial_number_device,
            "photo_ITAM_device": name_photo_ITAM_device,
        }
        res = posts.insert_one(post).inserted_id
    return {"message": "Запись успешно добавлена"}

@app.delete("/file/{name}")
async def delete_file(request: Request,name: str):
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postsSession.find_one({"Session": session_value})
    if(res==None):
        raise HTTPException(status_code=401, detail="Аккаунт не найден")
    path = "static/images/"
    path+=name
    try:
        if(path!="static/images/"):
            os.remove(path)
    except FileNotFoundError:
        pass
    return {"message": "Файл успешно удалён"}

@app.post("/signIn")
async def signIn(request: Request):
    data = await request.json()
    email = data["email"]
    password = data["password"]
    res = postsPassword.find_one({'email': email})
    if(res==None):
        raise HTTPException(status_code=401, detail="Аккаунт не найден")
    salt = res['salt']
    hash = res['password']
    newHash = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 100000).hex()
    if(hash != newHash):
        raise HTTPException(status_code=422, detail="Не правильный пароль")
    hashSession = hashlib.pbkdf2_hmac('sha256', email.encode('utf-8'),os.urandom(16).hex().encode('utf-8'), 100000).hex()
    content = {"message": "true"}
    response = JSONResponse(content=content)
    response.set_cookie(key="session", value=hashSession)

    try:
        filterDelete = {'id':res["id"]}
        postsSession.delete_one(filterDelete)
    except ValueError:
        pass

    n=postsSession.count_documents({})
    if( postsSession.count_documents({}) == n):
        post = {
            "id": res["id"],
            "Session": hashSession
        }
        res = postsSession.insert_one(post).inserted_id

    return response

@app.post("/signUp")
async def signUp(request: Request):
    data = await request.json()
    email = data["email"]
    password = data["password"]
    res = postsPassword.find_one({'email': email})
    if(res!=None):
        raise HTTPException(status_code=401, detail="Аккаунт не найден")
    if(len(password)<8):
        raise HTTPException(status_code=422, detail="Пароль слишком короткий")
    n=postsPassword.count_documents({})
    salt = os.urandom(16)
    hashPassword = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.hex().encode('utf-8'), 100000)
    newId = postsPassword.count_documents({})+1
    if( postsPassword.count_documents({}) == n):
        post = {
            "id": newId,
            "email": email,
            "password":hashPassword.hex(),
            "salt":salt.hex().encode('utf-8')
        }
        postsPassword.insert_one(post).inserted_id
    return {"message": "Регистрация успешна"}

@app.post("/exit")
async def auth(request:Request):
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postsSession.find_one({"Session": session_value})
    if(res==None):
        content = {"message": "true"}
        response = JSONResponse(content=content)
        response.delete_cookie(key="session")
        return response

    try:
        filterDelete = {'id':res["id"]}
        postsSession.delete_one(filterDelete)
    except ValueError:
        pass
    content = {"message": "true"}
    response = JSONResponse(content=content)
    response.delete_cookie(key="session")
    return response

@app.post("/export")
async def export(request:Request):
    cookies = request.cookies
    session_value = cookies.get("session")
    auth = postsSession.find_one({"Session": session_value})
    if(auth==None):
        raise HTTPException(status_code=401, detail="Аккаунт не найден")
    wb = Workbook()
    ws = wb.active

    res = posts.find({'user_id': auth["id"] })

    id = []
    type_device = []
    model_device = []
    serial_number = []
    ITAM_device = []
    photo_device = []
    photo_serial_number_device = []
    photo_ITAM_device = []
    files_to_add = []

    for document in res:
        id.append(document['id'])
        type_device.append(document['type_device'])
        model_device.append(document['model_device'])
        serial_number.append(document['serial_number'])
        ITAM_device.append(document['ITAM_device'])
        _photo_device = "static/images/"
        _photo_device+=document['photo_device']
        photo_device.append(_photo_device)
        files_to_add.append(_photo_device)
        _photo_serial_number_device = "static/images/"
        _photo_serial_number_device +=document['photo_serial_number_device']
        photo_serial_number_device.append(_photo_serial_number_device)
        files_to_add.append(_photo_serial_number_device)
        _photo_ITAM_device = "static/images/"
        _photo_ITAM_device +=document['photo_ITAM_device']
        photo_ITAM_device.append(_photo_ITAM_device)
        files_to_add.append(_photo_ITAM_device)

    with zipfile.ZipFile('Table Device.zip', 'w') as zip_file:
        pass


    wb.save('Table Device.xlsx')
    with zipfile.ZipFile('Table Device.zip', 'a') as zip_file:
        for file in files_to_add:
            zip_file.write(file)

    ws['A1'] = '№'
    ws['B1'] = 'Вид техники'
    ws['C1'] = 'Модель'
    ws['D1'] = 'Серийный номер'
    ws['E1'] = 'ITAM'
    ws['F1'] = 'Устройство'
    ws['G1'] = 'Серийный номер'
    ws['H1'] = 'Инвентарный номер'
    for i in range(len(id)):
        ws.cell(row=i+2, column=1).value = id[i]
        ws.cell(row=i+2, column=2).value = type_device[i]
        ws.cell(row=i+2, column=3).value = model_device[i]
        ws.cell(row=i+2, column=4).value = serial_number[i]
        ws.cell(row=i+2, column=5).value = ITAM_device[i]
        hyperlink_photo_device = Hyperlink(ref="", target=photo_device[i])
        ws.cell(row=i+2, column=6).hyperlink = hyperlink_photo_device
        hyperlink_photo_serial_number_device = Hyperlink(ref="", target=photo_serial_number_device[i])
        ws.cell(row=i+2, column=7).hyperlink = hyperlink_photo_serial_number_device
        hyperlink_photo_ITAM_device = Hyperlink(ref="", target=photo_ITAM_device[i])
        ws.cell(row=i+2, column=8).hyperlink = hyperlink_photo_ITAM_device

    for col_idx, col in enumerate(ws.columns, start=1):
        max_width = 0
        for cell in col:
            if cell.value:
                width = len(str(cell.value)) + 2
                if width > max_width:
                    max_width = width
        ws.column_dimensions[chr(64 + col_idx)].width = max_width

    wb.save('Table Device.xlsx')

    with zipfile.ZipFile('Table Device.zip', 'a') as zip_file:
        zip_file.write("Table Device.xlsx")

    return Response(
        content=open("Table Device.zip", "rb").read(),
        media_type="application/zip",
        headers={
            "Content-Disposition": "attachment; filename=Table Device.zip",
            "Cache-Control": "max-age=3600",
            "Accept-Encoding": "gzip, compress, br"
        }
    )

@app.get('/{name}/{size}')
async def test(request:Request,name: str,size: int):

    cookies = request.cookies
    session_value = cookies.get("session")
    auth = postsSession.find_one({"Session": session_value})
    if(auth==None):
        raise HTTPException(status_code=401, detail="Аккаунт не найден")

    img = Image.open(f"static/images/{name}")

    width_percent = (size / float(img.size[0]))

    height_size = int((float(img.size[1]) * float(width_percent)))

    new_image = img.resize((size, height_size))

    new_image.save("new_image.webp", format="webp")

    with open("new_image.webp", "rb") as f:
        image_data = f.read()

    os.remove("new_image.webp")

    return Response(content=image_data, media_type="image/webp", headers={"Cache-Control": "max-age=31536000"})
